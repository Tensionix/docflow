#!/usr/bin/env python3
"""
MD Tables -> XLSX

Extracts Markdown pipe tables (GFM style) from a .md file and writes them into an .xlsx workbook.
Each table becomes a separate worksheet: Table_1, Table_2, ...

Usage:
  python md_tables_to_xlsx.py --md report.md --out report.xlsx
"""

from __future__ import annotations

# >>> audion CLI bootstrap >>>
import os as _os, sys as _sys
_HERE = _os.path.dirname(_os.path.abspath(__file__))
if _HERE not in _sys.path:
    _sys.path.insert(0, _HERE)
try:
    _sys.stdout.reconfigure(encoding="utf-8")
    _sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass
# <<< audion CLI bootstrap <<<

import argparse
import re
from pathlib import Path

from openpyxl import Workbook

TABLE_ROW_RE = re.compile(r'^\s*\|.*\|\s*$')
SEP_RE = re.compile(r'^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*$')

def split_row(line: str) -> list[str]:
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    return [c.strip() for c in line.split("|")]

def extract_tables(md_text: str):
    lines = md_text.splitlines()
    tables = []
    i = 0
    while i < len(lines):
        if TABLE_ROW_RE.match(lines[i] or ""):
            header = split_row(lines[i])
            if i + 1 < len(lines) and SEP_RE.match(lines[i + 1] or ""):
                rows = []
                j = i + 2
                while j < len(lines) and TABLE_ROW_RE.match(lines[j] or ""):
                    rows.append(split_row(lines[j]))
                    j += 1
                tables.append((header, rows))
                i = j
                continue
        i += 1
    return tables

def main() -> int:
    ap = argparse.ArgumentParser(description="Convert Markdown pipe tables to XLSX.")
    ap.add_argument("--md", required=True, help="Input Markdown file")
    ap.add_argument("--out", required=True, help="Output XLSX file")
    args = ap.parse_args()

    in_md = Path(args.md).resolve()
    out_xlsx = Path(args.out).resolve()

    if not in_md.exists():
        print(f"[ERROR] Markdown file not found: {in_md}")
        return 2

    md_text = in_md.read_text(encoding="utf-8")
    tables = extract_tables(md_text)

    wb = Workbook()
    wb.remove(wb.active)

    if not tables:
        ws = wb.create_sheet("NoTables")
        ws["A1"] = "No Markdown pipe tables found."
        wb.save(out_xlsx)
        print(f"[OK] Wrote: {out_xlsx}")
        return 0

    for idx, (header, rows) in enumerate(tables, start=1):
        ws = wb.create_sheet(f"Table_{idx}")
        ws.append(header)
        for r in rows:
            if len(r) < len(header):
                r = r + [""] * (len(header) - len(r))
            elif len(r) > len(header):
                r = r[:len(header)]
            ws.append(r)

    wb.save(out_xlsx)
    print(f"[OK] Wrote: {out_xlsx} (tables: {len(tables)})")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
