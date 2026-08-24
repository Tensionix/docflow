#!/usr/bin/env python3
"""Remove non-meaningful Excel rows from XLSX copies."""

from __future__ import annotations

# >>> audion CLI bootstrap >>>
import os as _os, sys as _sys
_HERE = _os.path.dirname(_os.path.abspath(__file__))
if _HERE not in _sys.path:
    _sys.path.insert(0, _HERE)
try:
    _sys.stdout.reconfigure(encoding="utf-8")
    _sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass
# <<< audion CLI bootstrap <<<

import argparse
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from _office_common import md_escape, safe_mkdir, write_json_file


DEFAULT_SUFFIX = "_no_empty_rows"
SUPPORTED_SUFFIXES = {".xlsx"}


@dataclass
class CleanResult:
    source: Path
    output: Path
    status: str
    removed_rows: int = 0
    sheets: dict[str, int] | None = None
    error: str = ""


def has_meaningful_value(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def row_is_empty(worksheet, row_number: int) -> bool:
    values = (
        worksheet.cell(row=row_number, column=column).value
        for column in range(1, worksheet.max_column + 1)
    )
    return not any(has_meaningful_value(value) for value in values)


def _empty_row_runs(worksheet) -> list[tuple[int, int]]:
    runs: list[tuple[int, int]] = []
    run_start: int | None = None
    run_length = 0
    for row_number in range(1, worksheet.max_row + 1):
        if row_is_empty(worksheet, row_number):
            if run_start is None:
                run_start = row_number
                run_length = 0
            run_length += 1
            continue
        if run_start is not None:
            runs.append((run_start, run_length))
            run_start = None
            run_length = 0
    if run_start is not None:
        runs.append((run_start, run_length))
    return runs


def compact_worksheet(worksheet) -> int:
    runs = _empty_row_runs(worksheet)
    removed = sum(length for _, length in runs)
    for start, length in reversed(runs):
        worksheet.delete_rows(start, length)
    return removed


def clean_workbook(input_path: Path, output_path: Path, *, all_sheets: bool) -> dict[str, int]:
    workbook = load_workbook(input_path)
    try:
        worksheets = workbook.worksheets if all_sheets else [workbook.active]
        removed_by_sheet: dict[str, int] = {}
        for worksheet in worksheets:
            removed_by_sheet[worksheet.title] = compact_worksheet(worksheet)
        safe_mkdir(output_path.parent)
        workbook.save(output_path)
        return removed_by_sheet
    finally:
        workbook.close()


def output_path_for(input_path: Path, input_root: Path, out_dir: Path, suffix: str = DEFAULT_SUFFIX) -> Path:
    relative = input_path.name if input_path.is_file() and input_root == input_path.parent else input_path.relative_to(input_root)
    output = out_dir / relative
    return output.with_name(f"{output.stem}{suffix}{output.suffix}")


def iter_input_files(path: Path) -> tuple[Path, list[Path]]:
    resolved = path.resolve()
    if resolved.is_file():
        if resolved.suffix.lower() in SUPPORTED_SUFFIXES and not resolved.name.startswith("~$"):
            return resolved.parent, [resolved]
        return resolved.parent, []
    return resolved, sorted(
        file
        for file in resolved.rglob("*")
        if file.is_file() and file.suffix.lower() in SUPPORTED_SUFFIXES and not file.name.startswith("~$")
    )


def write_report(out_path: Path, input_root: Path, results: list[CleanResult], *, all_sheets: bool) -> None:
    safe_mkdir(out_path.parent)
    lines = ["# Очистка незначащих строк Excel\n"]
    lines.append(f"- Вход: `{md_escape(str(input_root))}`")
    lines.append(f"- Листы: **{'все' if all_sheets else 'только активный'}**")
    lines.append(f"- Файлов обработано: **{len(results)}**")
    lines.append(f"- Удалено строк: **{sum(item.removed_rows for item in results)}**")
    lines.append("")
    lines.append("| Файл | Статус | Удалено строк | Выходной файл | Детали |")
    lines.append("|---|---|---:|---|---|")
    for item in results:
        details = item.error
        if item.sheets:
            details = "; ".join(f"{name}: {count}" for name, count in item.sheets.items())
        lines.append(
            f"| `{md_escape(str(item.source))}` | `{item.status}` | {item.removed_rows} | `{md_escape(str(item.output))}` | {md_escape(details)} |"
        )
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_json(input_root: Path, results: list[CleanResult], *, all_sheets: bool) -> dict[str, object]:
    return {
        "tool": "xlsx_remove_empty_rows",
        "input_root": str(input_root),
        "all_sheets": all_sheets,
        "summary": {
            "files": len(results),
            "ok": sum(1 for item in results if item.status == "OK"),
            "failed": sum(1 for item in results if item.status != "OK"),
            "removed_rows": sum(item.removed_rows for item in results),
        },
        "files": [
            {
                "source": str(item.source),
                "output": str(item.output),
                "status": item.status,
                "removed_rows": item.removed_rows,
                "sheets": item.sheets or {},
                "error": item.error,
            }
            for item in results
        ],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Remove non-meaningful rows from XLSX copies.")
    parser.add_argument("--input", default="input", help="Input XLSX file or folder")
    parser.add_argument("--outdir", default="output/xlsx_no_empty_rows", help="Output folder")
    parser.add_argument("--report", default="report/xlsx_remove_empty_rows.md", help="Markdown report path")
    parser.add_argument("--json-out", default="", help="Optional JSON report path")
    parser.add_argument("--all-sheets", action="store_true", help="Clean all worksheets instead of only the active sheet")
    args = parser.parse_args()

    input_root, files = iter_input_files(Path(args.input))
    out_dir = Path(args.outdir).resolve()
    report_path = Path(args.report).resolve()
    results: list[CleanResult] = []

    if not input_root.exists():
        print(f"[ERROR] Input does not exist: {input_root}")
        return 2
    if not files:
        safe_mkdir(report_path.parent)
        report_path.write_text("# Очистка незначащих строк Excel\n\nФайлы XLSX не найдены.\n", encoding="utf-8")
        print(f"[WARN] No XLSX files found: {input_root}")
        print(f"[OK] Report: {report_path}")
        return 0

    for source in files:
        output = output_path_for(source, input_root, out_dir)
        try:
            removed_by_sheet = clean_workbook(source, output, all_sheets=args.all_sheets)
            removed = sum(removed_by_sheet.values())
            results.append(CleanResult(source=source, output=output, status="OK", removed_rows=removed, sheets=removed_by_sheet))
            print(f"[OK] {source} -> {output} removed={removed}")
        except Exception as exc:
            results.append(CleanResult(source=source, output=output, status="FAILED", error=str(exc)))
            print(f"[FAILED] {source}: {exc}")

    write_report(report_path, input_root, results, all_sheets=args.all_sheets)
    if args.json_out:
        json_path = Path(args.json_out).resolve()
        write_json_file(json_path, build_json(input_root, results, all_sheets=args.all_sheets))
        print(f"[OK] JSON: {json_path}")
    print(f"[OK] Report: {report_path}")
    return 1 if any(item.status != "OK" for item in results) else 0


if __name__ == "__main__":
    raise SystemExit(main())
